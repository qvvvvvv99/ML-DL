import openpyxl

# 데이터 변수 선언
filename="sample2.xlsx"

# 엑셀 파일 열기
book=openpyxl.load_workbook(filename)

# 시트 추출
sheet = book.worksheets[1]
print('sheet => ', sheet.title)
print('sheet.values => ', sheet.values)
print(f'sheet.max_row = {sheet.max_row}, sheet.max_colum={sheet.max_column}')
print(f'sheet.rows => {sheet.rows}')

# 시트에 행(row) 데이터 열기
rowdata=[]
for row in sheet.rows:
    print(f'row => {row}, {type(row)}, {len(row)}')
    print(f'row => {row[0].value}, {row[sheet.max_column-1].value}')
    rowdata.append([
        row[0].value,
        row[sheet.max_column-1].value
    ])
print('len(rowdata) => ', len(rowdata))