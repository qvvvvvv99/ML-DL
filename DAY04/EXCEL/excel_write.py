from openpyxl import Workbook

# 데이터 변수 선언
filename = 'sample.xlsx'

# 엑셀 파일 생성
wb = Workbook()         # 엑셀 파일 생성, Sheet1 자동 생성
ws = wb.active          # 시트 활성화
ws.title = 'new sheet'  # 시트명 변경
ws['A1'] = 'Language'   # A1 시트에 데이터('Language') 삽입
ws['B1'] = 'Create'

wb.save(filename=filename)